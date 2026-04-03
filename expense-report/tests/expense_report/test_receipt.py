from expense_report.receipt import attach_receipts, collect_receipt_files, validate_taxi_receipts
import openpyxl, os, tempfile

TEST_FOLDER = "/Users/ykh/Documents/drive/개인경비 지출결의서/2026/202603"


def test_collect_receipt_files():
    files = collect_receipt_files(TEST_FOLDER)
    assert len(files) >= 3
    assert all(os.path.splitext(f)[1].lower() in {".jpg", ".jpeg", ".png"} for f in files)


def test_collect_sorted_by_name():
    files = collect_receipt_files(TEST_FOLDER)
    names = [os.path.basename(f) for f in files]
    assert names == sorted(names)


def test_attach_receipts_adds_images():
    files = collect_receipt_files(TEST_FOLDER)
    with tempfile.TemporaryDirectory() as tmpdir:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "영수증 첨부"
        test_path = os.path.join(tmpdir, "test.xlsx")
        wb.save(test_path)
        wb = openpyxl.load_workbook(test_path)
        attach_receipts(wb["영수증 첨부"], files)
        wb.save(test_path)
        wb2 = openpyxl.load_workbook(test_path)
        assert len(wb2["영수증 첨부"]._images) == len(files)


def test_validate_taxi_receipts_warns():
    assert validate_taxi_receipts(True, []) is not None
    assert "택시비" in validate_taxi_receipts(True, [])


def test_validate_taxi_no_warning_with_receipts():
    assert validate_taxi_receipts(True, ["/some/file.jpg"]) is None


def test_validate_no_taxi_no_warning():
    assert validate_taxi_receipts(False, []) is None
