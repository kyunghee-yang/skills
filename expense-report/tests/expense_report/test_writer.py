from expense_report.classifier import Classification
from expense_report.parser import Transaction
from expense_report.writer import write_expense_report
import openpyxl, os, tempfile


def _txn(date, time, merchant, amount):
    return Transaction(date=date, time=time, merchant=merchant,
        card_number="4201-****-****-7592", usage_type="국내일반",
        amount=amount, transaction_type="국내 일시불",
        approval_number="12345678", purchase_status="매입",
        purchase_date="2026-03-28", installment="-", vat="1000", status="정상")


def test_write_creates_file():
    txns = [_txn("2026.03.27", "13:00", "바나프레소", 10300)]
    cls = [Classification(usage="팀 커피", expense_amount=10300, account="복리후생비[회식비]", companion="양경희,김보민", rule_number=1)]
    with tempfile.TemporaryDirectory() as d:
        out = os.path.join(d, "test.xlsx")
        write_expense_report(txns, cls, out)
        assert os.path.exists(out)


def test_sheet1_data_filled():
    txns = [_txn("2026.03.27", "13:00", "바나프레소", 10300)]
    cls = [Classification(usage="팀 커피", expense_amount=10300, account="복리후생비[회식비]", companion="양경희,김보민", rule_number=1)]
    with tempfile.TemporaryDirectory() as d:
        out = os.path.join(d, "test.xlsx")
        write_expense_report(txns, cls, out)
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["1.매출내역(원본)"]
        assert ws.cell(10, 1).value == "2026.03.27"
        assert ws.cell(10, 7).value == "바나프레소"
        assert ws.cell(10, 10).value == 10300


def test_sheet2_classification_filled():
    txns = [_txn("2026.03.27", "13:00", "바나프레소", 10300)]
    cls = [Classification(usage="팀 커피", expense_amount=10300, account="복리후생비[회식비]", companion="양경희,김보민", rule_number=1)]
    with tempfile.TemporaryDirectory() as d:
        out = os.path.join(d, "test.xlsx")
        write_expense_report(txns, cls, out)
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["2.(기명카드)사용내역"]
        assert ws.cell(6, 1).value == "양경희"
        assert ws.cell(6, 2).value == "R&D본부"
        assert ws.cell(6, 7).value == 10300
        assert ws.cell(6, 10).value == "팀 커피"
        assert ws.cell(6, 11).value == "양경희,김보민"
        assert ws.cell(6, 14).value == "복리후생비[회식비]"


def test_manual_items_leave_cells_empty():
    txns = [_txn("2026.03.27", "09:00", "네이버페이", 50000)]
    cls = [Classification(is_manual=True, rule_number=8, manual_fields=["usage", "expense_amount", "account", "companion"])]
    with tempfile.TemporaryDirectory() as d:
        out = os.path.join(d, "test.xlsx")
        write_expense_report(txns, cls, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["2.(기명카드)사용내역"]
        assert ws.cell(6, 1).value == "양경희"   # 기안자는 채워짐
        assert ws.cell(6, 10).value is None       # 사용내역 비어있음
        assert ws.cell(6, 14).value is None       # 계정과목 비어있음
