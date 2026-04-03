from __future__ import annotations
import copy
import warnings
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from expense_report.classifier import Classification
from expense_report.config import (
    DEPARTMENT, DRAFTER_NAME,
    SHEET1_COL_AMOUNT, SHEET1_COL_APPROVAL, SHEET1_COL_CARD,
    SHEET1_COL_DATE, SHEET1_COL_INSTALLMENT, SHEET1_COL_MERCHANT,
    SHEET1_COL_PURCHASE, SHEET1_COL_PURCHASE_DATE, SHEET1_COL_TIME,
    SHEET1_COL_TXN_TYPE, SHEET1_COL_TYPE, SHEET1_COL_VAT,
    SHEET1_DATA_START_ROW,
    SHEET2_COL_ACCOUNT, SHEET2_COL_COMPANION, SHEET2_COL_DEPT,
    SHEET2_COL_DRAFTER, SHEET2_COL_EXPENSE, SHEET2_COL_ROUTE,
    SHEET2_COL_USAGE, SHEET2_DATA_START_ROW,
    TEMPLATE_PATH,
)
from expense_report.parser import Transaction, XlsMeta

# 템플릿 기본 14쌍 (row 10~37)
_TEMPLATE_PAIR_COUNT = 14

# 각 2행 쌍의 병합 패턴 (col_start, col_end, row_span)
# row_span: 2 = 두 행 모두 병합, 1 = 첫 행만 (S열의 부가세/상태)
_MERGE_PATTERN = [
    (1, 3, 2),    # A:C
    (4, 6, 2),    # D:F
    (7, 7, 2),    # G
    (8, 8, 2),    # H
    (9, 9, 2),    # I
    (10, 10, 2),  # J
    (11, 11, 2),  # K
    (12, 14, 2),  # L:N
    (15, 15, 2),  # O
    (16, 17, 2),  # P:Q
    (18, 18, 2),  # R
    (19, 21, 1),  # S:U (첫 행만 — 둘째 행은 상태)
]


def _write_sheet1_meta(ws, meta: XlsMeta) -> None:
    ws.cell(3, 5).value = meta.period
    ws.cell(4, 6).value = meta.domestic_count
    ws.cell(4, 14).value = meta.domestic_total
    ws.cell(5, 6).value = meta.overseas_count
    ws.cell(5, 14).value = meta.overseas_total
    ws.cell(6, 6).value = meta.cancel_count
    ws.cell(6, 14).value = meta.reject_count


def _copy_cell_style(src_cell, dst_cell) -> None:
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.protection = copy.copy(src_cell.protection)


def _ensure_pairs(ws, needed: int) -> None:
    """데이터 쌍이 부족하면 템플릿 row 10-11의 스타일/병합을 복제하여 확장."""
    if needed <= _TEMPLATE_PAIR_COUNT:
        return

    # footer 병합 제거 (나중에 다시 추가)
    footer_merges = [
        str(mr) for mr in list(ws.merged_cells.ranges)
        if mr.min_row >= SHEET1_DATA_START_ROW + _TEMPLATE_PAIR_COUNT * 2
    ]
    for fm in footer_merges:
        ws.unmerge_cells(fm)

    # 새 쌍 생성 (기존 14쌍 이후부터)
    template_row = SHEET1_DATA_START_ROW  # row 10 기준 스타일 복제
    max_col = 21  # U열

    for pair_idx in range(_TEMPLATE_PAIR_COUNT, needed):
        new_row = SHEET1_DATA_START_ROW + pair_idx * 2

        # 스타일 복제 (2행)
        for offset in range(2):
            src_row = template_row + offset
            dst_row = new_row + offset
            for col in range(1, max_col + 1):
                src_cell = ws.cell(src_row, col)
                dst_cell = ws.cell(dst_row, col)
                _copy_cell_style(src_cell, dst_cell)

        # 병합 생성
        for col_start, col_end, row_span in _MERGE_PATTERN:
            if col_start == col_end and row_span == 2:
                # 단일 열 2행 병합
                merge_range = f"{get_column_letter(col_start)}{new_row}:{get_column_letter(col_end)}{new_row + 1}"
            elif row_span == 1:
                # 첫 행만 병합 (S:U)
                merge_range = f"{get_column_letter(col_start)}{new_row}:{get_column_letter(col_end)}{new_row}"
            else:
                # 다중 열 2행 병합
                merge_range = f"{get_column_letter(col_start)}{new_row}:{get_column_letter(col_end)}{new_row + 1}"
            ws.merge_cells(merge_range)


def _write_sheet1(ws, all_transactions: list[Transaction]) -> None:
    needed_pairs = len(all_transactions)
    _ensure_pairs(ws, needed_pairs)

    row = SHEET1_DATA_START_ROW
    for txn in all_transactions:
        ws.cell(row, SHEET1_COL_DATE).value = txn.date
        ws.cell(row, SHEET1_COL_TIME).value = txn.time
        ws.cell(row, SHEET1_COL_MERCHANT).value = txn.merchant
        ws.cell(row, SHEET1_COL_CARD).value = txn.card_number
        ws.cell(row, SHEET1_COL_TYPE).value = txn.usage_type
        ws.cell(row, SHEET1_COL_AMOUNT).value = txn.amount
        ws.cell(row, SHEET1_COL_TXN_TYPE).value = txn.transaction_type
        ws.cell(row, SHEET1_COL_APPROVAL).value = txn.approval_number
        ws.cell(row, SHEET1_COL_PURCHASE).value = txn.purchase_status
        ws.cell(row, SHEET1_COL_PURCHASE_DATE).value = txn.purchase_date
        ws.cell(row, SHEET1_COL_INSTALLMENT).value = txn.installment
        ws.cell(row, SHEET1_COL_VAT).value = txn.vat
        ws.cell(row + 1, SHEET1_COL_VAT).value = txn.status
        row += 2


def _write_sheet2(ws, transactions: list[Transaction], classifications: list[Classification]) -> None:
    row = SHEET2_DATA_START_ROW
    for txn, cls in zip(transactions, classifications):
        ws.cell(row, SHEET2_COL_DRAFTER).value = DRAFTER_NAME
        ws.cell(row, SHEET2_COL_DEPT).value = DEPARTMENT
        if not cls.is_manual or "expense_amount" not in cls.manual_fields:
            ws.cell(row, SHEET2_COL_EXPENSE).value = cls.expense_amount
        if cls.usage:
            ws.cell(row, SHEET2_COL_USAGE).value = cls.usage
        if cls.companion:
            ws.cell(row, SHEET2_COL_COMPANION).value = cls.companion
        if cls.route:
            ws.cell(row, SHEET2_COL_ROUTE).value = cls.route
        if cls.account:
            ws.cell(row, SHEET2_COL_ACCOUNT).value = cls.account
        row += 1


def write_expense_report(
    transactions: list[Transaction],
    classifications: list[Classification],
    output_path: str,
    all_transactions: Optional[list[Transaction]] = None,
    meta: Optional[XlsMeta] = None,
) -> None:
    warnings.filterwarnings("ignore", category=UserWarning)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    sheet1 = wb["1.매출내역(원본)"]
    if meta:
        _write_sheet1_meta(sheet1, meta)
    _write_sheet1(sheet1, all_transactions or transactions)
    _write_sheet2(wb["2.(기명카드)사용내역"], transactions, classifications)
    wb.save(output_path)
