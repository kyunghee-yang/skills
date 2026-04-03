from __future__ import annotations
import json
import os
import sys
import warnings
from collections import Counter
from typing import Optional

import openpyxl

from expense_report.classifier import classify, Classification
from expense_report.config import (
    OUTPUT_FILENAME_TEMPLATE,
    DRAFTER_NAME,
    SHEET2_COL_ACCOUNT, SHEET2_COL_COMPANION, SHEET2_COL_EXPENSE,
    SHEET2_COL_USAGE, SHEET2_DATA_START_ROW,
)
from expense_report.matcher import NotionEntry, match_transactions
from expense_report.parser import parse_xls, parse_xls_all, parse_xls_meta
from expense_report.receipt import attach_receipts, collect_receipt_files, validate_taxi_receipts
from expense_report.writer import write_expense_report


def _extract_year_month(folder_path: str) -> tuple:
    basename = os.path.basename(folder_path.rstrip("/"))
    if len(basename) == 6 and basename.isdigit():
        return basename[2:4], basename[4:6]
    raise ValueError(f"폴더명에서 연월을 추출할 수 없습니다: {basename}")


def _find_xls(folder_path: str) -> str:
    for fname in os.listdir(folder_path):
        if fname.endswith(".xls") and not fname.startswith(".") and not fname.endswith(".xlsx"):
            return os.path.join(folder_path, fname)
    raise FileNotFoundError(f"간편서비스_승인내역.xls 파일을 찾을 수 없습니다: {folder_path}")


def _read_existing_overrides(output_path: str) -> dict[int, dict]:
    """기존 출력 파일에서 수기 입력값을 읽어 보존할 항목 반환.

    Returns: {row_index: {"account": str, "usage": str, "expense": int, "companion": str}}
    """
    if not os.path.exists(output_path):
        return {}

    warnings.filterwarnings("ignore", category=UserWarning)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["2.(기명카드)사용내역"]

    overrides = {}
    for row in range(SHEET2_DATA_START_ROW, ws.max_row + 1):
        account = ws.cell(row, SHEET2_COL_ACCOUNT).value
        if not account:
            continue
        idx = row - SHEET2_DATA_START_ROW
        overrides[idx] = {
            "account": account,
            "usage": ws.cell(row, SHEET2_COL_USAGE).value,
            "expense": ws.cell(row, SHEET2_COL_EXPENSE).value,
            "companion": ws.cell(row, SHEET2_COL_COMPANION).value,
        }
    wb.close()
    return overrides


def run_pipeline(folder_path: str, notion_data: Optional[dict] = None) -> dict:
    yy, mm = _extract_year_month(folder_path)
    xls_path = _find_xls(folder_path)
    all_transactions = parse_xls_all(xls_path)
    meta = parse_xls_meta(xls_path)

    # 정상 거래만 추출 (Notion 매칭용)
    normal_transactions = [t for t in all_transactions if t.status != "취소"]

    # Notion matching (정상 거래만 대상)
    notion_entries = []
    if notion_data and "entries" in notion_data:
        for entry_dict in notion_data["entries"]:
            notion_entries.append(NotionEntry(
                date=entry_dict["date"],
                amount=entry_dict["amount"],
                companions=entry_dict["companions"],
            ))
    normal_matches = match_transactions(normal_transactions, notion_entries)

    # 정상 거래 index → all_transactions index 매핑
    normal_idx_to_all = {}
    ni = 0
    for ai, txn in enumerate(all_transactions):
        if txn.status != "취소":
            normal_idx_to_all[ni] = ai
            ni += 1

    # all_transactions index → Notion match 역매핑
    all_matches = {}
    for ni, notion_entry in normal_matches.items():
        all_matches[normal_idx_to_all[ni]] = notion_entry

    # Classify ALL transactions (Sheet1과 Sheet2 행 수 일치)
    output_filename = OUTPUT_FILENAME_TEMPLATE.format(yy=yy, mm=mm)
    output_path = os.path.join(folder_path, output_filename)

    classifications = []
    for idx, txn in enumerate(all_transactions):
        if txn.status == "취소":
            cls = classify(txn, notion_match=None)
            cls.expense_amount = 0
        else:
            cls = classify(txn, notion_match=all_matches.get(idx))
        classifications.append(cls)

    # 기존 출력 파일에서 수기 입력 보존 (Rule 8 항목만)
    existing_overrides = _read_existing_overrides(output_path)
    for idx, cls in enumerate(classifications):
        if cls.rule_number != 8:
            continue
        override = existing_overrides.get(idx)
        if not override or not override["account"]:
            continue
        classifications[idx] = Classification(
            usage=override.get("usage") or "",
            expense_amount=override.get("expense") or all_transactions[idx].amount,
            account=override["account"],
            companion=override.get("companion") or DRAFTER_NAME,
            rule_number=0,
        )
    write_expense_report(
        all_transactions, classifications, output_path,
        all_transactions=all_transactions, meta=meta,
    )

    # Attach receipts
    receipt_files = collect_receipt_files(folder_path)
    has_taxi = any(c.rule_number in (2, 3, 4) for c in classifications)
    taxi_warning = validate_taxi_receipts(has_taxi, receipt_files)

    if receipt_files:
        import openpyxl
        warnings.filterwarnings("ignore", category=UserWarning)
        wb = openpyxl.load_workbook(output_path)
        attach_receipts(wb["영수증 첨부"], receipt_files)
        wb.save(output_path)

    # Summary
    rule_counter = Counter(c.rule_number for c in classifications)
    manual_items = []
    for idx, (txn, cls) in enumerate(zip(all_transactions, classifications)):
        if cls.rule_number == 8 and txn.status != "취소":
            manual_items.append({
                "row": idx + 1,
                "date": txn.date,
                "time": txn.time,
                "merchant": txn.merchant,
                "amount": txn.amount,
            })

    return {
        "created_file": output_path,
        "total_count": len(all_transactions),
        "classified_summary": {f"rule_{k}": v for k, v in sorted(rule_counter.items())},
        "manual_items": manual_items,
        "manual_count": len(manual_items),
        "taxi_receipt_warning": taxi_warning,
        "receipt_count": len(receipt_files),
    }


def main():
    import argparse
    ap = argparse.ArgumentParser(description="기명카드 지출결의서 생성")
    ap.add_argument("--folder", required=True, help="대상 폴더 경로")
    ap.add_argument("--notion-data", help="Notion 매칭 데이터 JSON 파일 경로")
    args = ap.parse_args()

    notion_data = None
    if args.notion_data:
        with open(args.notion_data) as f:
            notion_data = json.load(f)

    result = run_pipeline(args.folder, notion_data)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
